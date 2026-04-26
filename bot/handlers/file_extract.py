"""Извлечение текста из популярных форматов документов.

Каждая функция работает в формате «попробуй и верни текст или None». Все
зависимости pure-Python чтобы работало на Termux/Android без сборки C-расширений:

- **PDF**  — `pypdf` (бывший PyPDF2). Работает с обычными PDF; для PDF из
  сканов (картинки) текста нет — функция вернёт пустую строку, что мы
  передадим как «не удалось извлечь текст».
- **DOCX** — `docx2txt`. Параграфы + таблицы + базовое форматирование.
- **XLSX/XLSM** — `openpyxl`. Каждый лист → CSV-подобный дамп.
- **RTF**  — `striprtf`. Простой текст из RTF.
- **CSV/TSV/HTML/прочий plain-text** обрабатывается отдельно в files.py
  (через декодирование байтов как text).

Все экстракторы ловят свои исключения и возвращают `None` — caller сам
решит как сообщить пользователю.
"""

from __future__ import annotations

import io
import logging
import re
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class ExtractedDoc:
    """Результат извлечения. На случай ошибки заполняется error."""

    text: str
    pages_or_sheets: int = 0  # сколько страниц/листов прочитали (для UX)
    format: str = ""  # "pdf", "docx", и т.п.
    error: str | None = None


# Расширения, которые мы умеем расшифровать структурированно.
SUPPORTED_DOC_EXTS: set[str] = {
    ".pdf", ".docx", ".doc", ".rtf",
    ".xlsx", ".xlsm", ".xltx",
    ".odt",  # пробуем как docx через zip+xml — иногда работает
}


def extract_pdf(path: Path) -> ExtractedDoc:
    try:
        from pypdf import PdfReader
    except Exception as exc:  # noqa: BLE001
        return ExtractedDoc("", format="pdf", error=f"pypdf не установлен: {exc}")
    try:
        reader = PdfReader(str(path))
        if reader.is_encrypted:
            try:
                reader.decrypt("")  # пытаемся открыть с пустым паролем
            except Exception:  # noqa: BLE001
                return ExtractedDoc("", format="pdf", error="PDF зашифрован паролем")
        parts: list[str] = []
        for i, page in enumerate(reader.pages, start=1):
            try:
                t = page.extract_text() or ""
            except Exception as exc:  # noqa: BLE001
                logger.warning("pdf page %d failed: %s", i, exc)
                continue
            t = t.strip()
            if t:
                parts.append(f"--- Страница {i} ---\n{t}")
        text = "\n\n".join(parts).strip()
        if not text:
            return ExtractedDoc(
                "",
                pages_or_sheets=len(reader.pages),
                format="pdf",
                error="PDF не содержит текста (возможно скан — нужно OCR)",
            )
        return ExtractedDoc(text=text, pages_or_sheets=len(reader.pages), format="pdf")
    except Exception as exc:  # noqa: BLE001
        logger.exception("PDF extract failed")
        return ExtractedDoc("", format="pdf", error=f"не смог прочитать PDF: {exc}")


def extract_docx(path: Path) -> ExtractedDoc:
    try:
        import docx2txt
    except Exception as exc:  # noqa: BLE001
        return ExtractedDoc("", format="docx", error=f"docx2txt не установлен: {exc}")
    try:
        text = docx2txt.process(str(path)) or ""
        text = re.sub(r"\n{3,}", "\n\n", text).strip()
        if not text:
            return ExtractedDoc("", format="docx", error="документ пуст")
        # docx2txt не сообщает количество страниц; используем количество
        # параграфов как грубую метрику.
        paragraphs = sum(1 for line in text.splitlines() if line.strip())
        return ExtractedDoc(text=text, pages_or_sheets=paragraphs, format="docx")
    except Exception as exc:  # noqa: BLE001
        logger.exception("DOCX extract failed")
        return ExtractedDoc("", format="docx", error=f"не смог прочитать .docx: {exc}")


def extract_xlsx(path: Path, *, max_rows_per_sheet: int = 200) -> ExtractedDoc:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        return ExtractedDoc("", format="xlsx", error=f"openpyxl не установлен: {exc}")
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.exception("XLSX open failed")
        return ExtractedDoc("", format="xlsx", error=f"не смог открыть .xlsx: {exc}")
    # Гарантируем закрытие воркбука даже при ошибке итерации — иначе
    # openpyxl в read_only режиме держит ZIP-handle открытым.
    try:
        parts: list[str] = []
        for ws in wb.worksheets:
            rows: list[str] = [f"--- Лист «{ws.title}» ---"]
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if r_idx > max_rows_per_sheet:
                    # ws.max_row может быть None в read_only режиме,
                    # если в файле нет dimension-метаданных.
                    remaining = (ws.max_row or 0) - max_rows_per_sheet
                    if remaining > 0:
                        rows.append(f"... (ещё ~{remaining} строк)")
                    else:
                        rows.append("... (обрезано)")
                    break
                cells = [
                    "" if v is None else str(v).replace("\n", " ").replace("\t", " ")
                    for v in row
                ]
                # Пропускаем абсолютно пустые строки.
                if not any(c.strip() for c in cells):
                    continue
                rows.append("\t".join(cells))
            parts.append("\n".join(rows))
        text = "\n\n".join(parts).strip()
        if not text:
            return ExtractedDoc("", format="xlsx", error="книга пуста")
        return ExtractedDoc(text=text, pages_or_sheets=len(parts), format="xlsx")
    except Exception as exc:  # noqa: BLE001
        logger.exception("XLSX extract failed")
        return ExtractedDoc("", format="xlsx", error=f"не смог прочитать .xlsx: {exc}")
    finally:
        with suppress(Exception):
            wb.close()


def extract_rtf(path: Path) -> ExtractedDoc:
    try:
        from striprtf.striprtf import rtf_to_text
    except Exception as exc:  # noqa: BLE001
        return ExtractedDoc("", format="rtf", error=f"striprtf не установлен: {exc}")
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
        text = rtf_to_text(raw, errors="ignore").strip()
        if not text:
            return ExtractedDoc("", format="rtf", error="документ пуст")
        return ExtractedDoc(text=text, pages_or_sheets=1, format="rtf")
    except Exception as exc:  # noqa: BLE001
        logger.exception("RTF extract failed")
        return ExtractedDoc("", format="rtf", error=f"не смог прочитать .rtf: {exc}")


def extract_doc(path: Path) -> ExtractedDoc:
    """Старый бинарный .doc (Word 97-2003).

    Pure-Python библиотек надёжных нет. Возвращаем понятную ошибку, чтобы
    пользователь пересохранил в .docx или .pdf.
    """
    return ExtractedDoc(
        "",
        format="doc",
        error=(
            "формат .doc (Word 97-2003) не поддерживается. "
            "Сохраните файл как .docx или .pdf и пришлите снова."
        ),
    )


def render_pdf_to_jpegs(
    path: Path,
    *,
    max_pages: int = 8,
    scale: float = 1.5,
    jpeg_quality: int = 85,
) -> tuple[list[bytes], int]:
    """Рендерит первые `max_pages` страниц PDF в JPEG-байты.

    Используется для OCR PDF-сканов: pypdf не достал текст → рендерим
    страницы в картинки и отдаём их vision-модели.

    Возвращает (список JPEG bytes, всего страниц в PDF).
    На ошибке (нет pypdfium2 / повреждённый PDF) — `([], 0)`.
    """
    try:
        import pypdfium2 as pdfium
    except Exception as exc:  # noqa: BLE001
        logger.warning("pypdfium2 не установлен: %s", exc)
        return [], 0
    try:
        doc = pdfium.PdfDocument(str(path))
    except Exception as exc:  # noqa: BLE001
        logger.warning("PDFium не смог открыть %s: %s", path, exc)
        return [], 0
    total = len(doc)
    out: list[bytes] = []
    try:
        for i in range(min(total, max_pages)):
            try:
                page = doc[i]
                pil = page.render(scale=scale).to_pil()
                # Стандартизуем в RGB JPEG, чтобы передавать в OpenAI vision.
                if pil.mode != "RGB":
                    pil = pil.convert("RGB")
                buf = io.BytesIO()
                pil.save(buf, format="JPEG", quality=jpeg_quality, optimize=True)
                out.append(buf.getvalue())
            except Exception as exc:  # noqa: BLE001
                logger.warning("PDF page %d render failed: %s", i + 1, exc)
                continue
    finally:
        with suppress(Exception):
            doc.close()
    return out, total


def extract_doc_by_ext(path: Path, ext: str) -> ExtractedDoc | None:
    """Диспетчер по расширению. None — формат не структурированный документ.

    Сюда передавать ТОЛЬКО расширения из SUPPORTED_DOC_EXTS — иначе функция
    просто вернёт ExtractedDoc с ошибкой «формат не поддержан».
    """
    ext = ext.lower()
    if ext == ".pdf":
        return extract_pdf(path)
    if ext in (".docx", ".odt"):
        return extract_docx(path)
    if ext in (".xlsx", ".xlsm", ".xltx"):
        return extract_xlsx(path)
    if ext == ".rtf":
        return extract_rtf(path)
    if ext == ".doc":
        return extract_doc(path)
    return None
